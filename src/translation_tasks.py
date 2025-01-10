import asyncio
from openpyxl.utils import get_column_letter
from progress import save_progress
from typing import List, Dict, Any
import pandas as pd
from config import Config
from translator import Translator

async def translate_cells(translator: Translator, sheet_name: str, column: str, rows: List[int],
                          workbook, progress_data: Dict[str, Any], config: Config):
    texts = []
    cell_refs = []
    for row in rows:
        cell = workbook[sheet_name][f"{column}{row}"]
        text = cell.value
        if text and not pd.isnull(text):
            texts.append(text)
            cell_refs.append(cell.coordinate)
        else:
            texts.append("")
            cell_refs.append(cell.coordinate)

    translations = await asyncio.gather(*[
        translator.translate_text(text, sheet_name) for text in texts
    ], return_exceptions=True)

    successful_translations = 0

    for coord, translation in zip(cell_refs, translations):
        if isinstance(translation, Exception):
            print(f"Translation failed for cell {coord}: {translation}")
            continue
            
        # Special handling for specific columns
        columns_to_format = {'npc_name', 'topic', 'knowledge_class_name'}
        if column.lower() in columns_to_format and translation:
            try:
                # For npc_name, remove notes using language-specific pattern
                if column.lower() == 'npc_name':
                    note_pattern = config.get_note_pattern()
                    if note_pattern:
                        # Handle both "Note:" and "(Note:" patterns
                        cleaned = str(translation).split(f'({note_pattern}')[0].split(note_pattern)[0].strip()
                    else:
                        cleaned = str(translation).strip()
                else:
                    cleaned = str(translation).strip()
                # Apply formatting to all specified columns
                cleaned = cleaned.replace("_", " ").title()
                if cleaned:
                    translation = cleaned
                else:
                    print(f"Warning: Empty result after cleaning {column} at {coord}")
            except Exception as e:
                print(f"Error cleaning {column} at {coord}: {e}")

            
        workbook[sheet_name][coord].value = translation
        successful_translations += 1

    progress_data['successful_translations'] += successful_translations
    progress_data['current_indices'][sheet_name][column] += len(rows)

    # Check if we've reached or passed the next save point
    if progress_data['successful_translations'] >= progress_data['next_save_at']:
        save_progress(progress_data, config, workbook)
        # Set next save point by adding save_interval to current translations
        progress_data['next_save_at'] = progress_data['successful_translations'] + config.save_interval

    return successful_translations

async def translate_sheet(translator: Translator, sheet_name: str, workbook, progress_data: Dict[str, Any], config: Config):
    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    max_col = sheet.max_column

    columns = []
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        cell = sheet[f"{col_letter}1"]
        if cell.value and cell.value in config.blocked_columns:
            print(f"Skipping blocked column '{cell.value}' in sheet '{sheet_name}'.")
            continue
        columns.append(col_letter)

    tasks = []
    batch_size = 50  # Number of rows per batch

    for col in columns:
        if sheet_name not in progress_data['current_indices']:
            progress_data['current_indices'][sheet_name] = {}
        if col not in progress_data['current_indices'][sheet_name]:
            progress_data['current_indices'][sheet_name][col] = 2  # Assuming first row is header

        current_row = progress_data['current_indices'][sheet_name][col]
        while current_row <= max_row:
            end_row = min(current_row + batch_size - 1, max_row)
            rows = list(range(current_row, end_row + 1))
            tasks.append(
                translate_cells(translator, sheet_name, col, rows, workbook, progress_data, config)
            )
            current_row = end_row + 1

    translations_done = 0
    for task in asyncio.as_completed(tasks):
        try:
            success = await task
            translations_done += success
        except Exception as e:
            print(f"Error during translation task: {e}")

    # After completing the sheet, save the file
    if sheet_name not in progress_data['completed_sheets']:
        save_progress(progress_data, config, workbook)
        progress_data['completed_sheets'].append(sheet_name)

    print(f"Completed translations for sheet '{sheet_name}'. Total successful translations: {translations_done}")
