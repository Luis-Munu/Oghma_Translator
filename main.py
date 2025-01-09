import logging
import asyncio
from asyncio import Semaphore
from config import Config
from translator import Translator
from progress import load_progress, save_progress
from openpyxl import load_workbook
from translation_tasks import translate_sheet
from openai import AsyncOpenAI

config = Config()

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

async def translate_all_sheets(translator: Translator):
    print("Loading workbook...")
    workbook = load_workbook(filename=config.input_file, read_only=False)

    progress = load_progress(config)

    for sheet_name in workbook.sheetnames:
        if sheet_name in progress['completed_sheets']:
            print(f"Skipping already completed sheet: {sheet_name}")
            continue

        print(f"Processing sheet: {sheet_name}")
        await translate_sheet(translator, sheet_name, workbook, progress)

    # Final save after all sheets are translated
    save_progress(progress, config, workbook)

    print(f"All translations completed and saved to '{config.output_file}'.")

async def main():
    try:
        semaphore = Semaphore(config.max_concurrent_requests)
        client = AsyncOpenAI(api_key=config.MODEL_API_KEY, base_url=config.base_url, timeout=config.timeout)
        translator = Translator(client, semaphore)
        await translate_all_sheets(translator)
    except KeyboardInterrupt:
        print("Translation interrupted by user.")
    except Exception as e:
        print(f"Fatal error during translation: {e}")
        print("Translation process completed and resources cleaned up.")

if __name__ == "__main__":
    asyncio.run(main())