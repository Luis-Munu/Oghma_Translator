import logging
import asyncio
from asyncio import Semaphore
from openpyxl import load_workbook
from openai import AsyncOpenAI
from translation_tasks import translate_sheet
from config import Config, supported_languages
from translator import Translator
from progress import load_progress, save_progress

async def translate_all_sheets(translator: Translator, config: Config):
    print(f"Loading workbook for {config.language}...")
    workbook = load_workbook(filename=config.input_file, read_only=False)
    progress = load_progress(config)

    for sheet_name in workbook.sheetnames:
        if sheet_name in progress['completed_sheets']:
            print(f"Skipping already completed sheet: {sheet_name}")
            continue

        print(f"Processing sheet: {sheet_name}")
        await translate_sheet(translator, sheet_name, workbook, progress, config)

    save_progress(progress, config, workbook)
    print(f"All translations completed and saved to '{config.output_file}'.")

async def main():
    # Initialize variables outside try block for exception handling
    config = None
    workbook = None
    progress = None
    
    try:
        # Create a default config instance for initializing resources
        default_config = Config(language='es')  # Default language for initialization
        semaphore = Semaphore(default_config.max_concurrent_requests)
        client = AsyncOpenAI(api_key=default_config.MODEL_API_KEY, base_url=default_config.base_url, timeout=default_config.timeout)

        for lang in supported_languages:
            config = Config(language=lang)
            workbook = load_workbook(filename=config.input_file, read_only=False)
            progress = load_progress(config)
            translator = Translator(client, semaphore, config)
            await translate_all_sheets(translator, config)
    except KeyboardInterrupt:
        print("\nTranslation interrupted by user. Saving current progress...")
        try:
            save_progress(progress, config, workbook)
            print(f"Progress saved at {progress['successful_translations']} translations.")
        except Exception as save_error:
            print(f"Error saving progress: {save_error}")
        print("Translation process interrupted. Partial results saved.")
    except Exception as e:
        print(f"Fatal error during translation: {e}")
        try:
            save_progress(progress, config, workbook)
            print(f"Progress saved at {progress['successful_translations']} translations before exiting.")
        except Exception as save_error:
            print(f"Error saving progress: {save_error}")
        print("Translation process completed with errors. Resources cleaned up.")

if __name__ == "__main__":
    asyncio.run(main())
